import json
import sys
import os
from datetime import datetime, timezone

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print("openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


def load_json(input_path: str) -> dict:
    with open(input_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def safe_get(dct, *keys, default=None):
    cur = dct
    for key in keys:
        if cur is None:
            return default
        if isinstance(cur, dict) and key in cur:
            cur = cur[key]
        else:
            return default
    return cur


def to_excel_datetime(value):
    if not value:
        return None
    if isinstance(value, (int, float)):
        try:
            # Interpret numeric as UTC epoch (ms or s) and return naive UTC
            seconds = value / 1000.0 if value > 1e12 else value
            return datetime.utcfromtimestamp(seconds)
        except Exception:
            return value
    if isinstance(value, str):
        # Try ISO 8601
        try:
            dt = datetime.fromisoformat(value.replace('Z', '+00:00'))
            # Convert to naive UTC for Excel
            if dt.tzinfo is not None:
                dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
            return dt
        except Exception:
            return value
    return value


def autosize_columns(ws):
    for column_cells in ws.columns:
        max_length = 0
        column = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                cell_val = cell.value
                if isinstance(cell_val, (dict, list)):
                    cell_val = json.dumps(cell_val, ensure_ascii=False)
                if cell_val is None:
                    length = 0
                else:
                    length = len(str(cell_val))
                max_length = max(max_length, length)
            except Exception:
                pass
        ws.column_dimensions[column].width = min(max(12, max_length + 2), 80)


def normalize_student_id(student_id):
    if not student_id:
        return None
    s = str(student_id).strip()
    # Match digits and optional -suffix digits, keep base
    import re
    m = re.match(r"^(\d+)(?:-\d+)?$", s)
    if m:
        return m.group(1)
    return s


def fingerprint_summary(fp: dict):
    if not isinstance(fp, dict):
        return None
    ua = fp.get("userAgent")
    res = fp.get("screenResolution")
    tz = fp.get("timezone") or fp.get("timeZone")
    parts = []
    if ua:
        parts.append(ua)
    if res:
        parts.append(res)
    if tz:
        parts.append(str(tz))
    return " | ".join(parts) if parts else None


def compute_answer_flags(ans: dict):
    ba = ans.get("behaviorAnalytics") or {}
    cm = ans.get("comprehensiveMetrics") or {}
    ai = cm.get("academicIntegrityMetrics") or {}
    attention = ai.get("attentionMetrics") or {}
    interface = cm.get("interfaceUsage") or {}

    student_answer = ans.get("studentAnswer") or ""
    typing_events = ans.get("typingEvents") or cm.get("keystrokeEvents") or []
    time_spent = ans.get("timeSpent")

    words_per_minute = ba.get("wordsPerMinute")
    average_key_interval = ba.get("averageKeyInterval")
    copy_paste_events = ba.get("copyPasteEvents") or 0

    suspicious_typing_speed = bool(ba.get("suspiciousTypingSpeed"))
    paste_reported = bool(ba.get("pasteFromExternal")) or (copy_paste_events > 0)
    devtools_opened = bool(ba.get("devToolsOpened") or interface.get("devToolsOpened"))
    tab_switches = ba.get("tabSwitches") or attention.get("tabSwitches") or 0
    window_blur = ba.get("windowBlurEvents") or attention.get("windowBlurEvents") or 0

    # Heuristic paste detection: long answer with few typing events, or very short time
    suspected_paste_heuristic = False
    try:
        if len(student_answer) >= 30 and len(typing_events) <= 3:
            suspected_paste_heuristic = True
        if time_spent is not None and time_spent <= 5 and len(student_answer) >= 10:
            suspected_paste_heuristic = True
    except Exception:
        pass

    high_wpm = False
    try:
        if words_per_minute is not None and float(words_per_minute) >= 120.0:
            high_wpm = True
    except Exception:
        pass

    return {
        "suspiciousTypingSpeed": suspicious_typing_speed,
        "pasteReported": paste_reported,
        "suspectedPasteHeuristic": suspected_paste_heuristic,
        "devToolsOpened": devtools_opened,
        "tabSwitches": tab_switches,
        "windowBlurEvents": window_blur,
        "wordsPerMinute": words_per_minute,
        "averageKeyInterval": average_key_interval,
        "copyPasteEvents": copy_paste_events,
        "typingEventsCount": len(typing_events),
        "answerLength": len(student_answer),
        "highWPM": high_wpm,
    }


def analyze_report(report: list):
    # Aggregates
    students = {}
    ip_to_students = {}

    # First pass: collect per-student and per-session stats
    for student in report:
        student_name = student.get("studentName")
        for entry in student.get("sessions", []):
            session = entry.get("session", {})
            answers = entry.get("answers", [])

            norm_id = normalize_student_id(session.get("studentId"))
            student_key = norm_id or student_name or session.get("studentEmail")

            agg = students.setdefault(student_key, {
                "studentKey": student_key,
                "names": set(),
                "normalizedIds": set(),
                "emails": set(),
                "sessions": [],
                "ips": [],  # preserve order occurrences
                "ips_set": set(),
                "userAgents": set(),
                "answersCount": 0,
                "devToolsAnswers": 0,
                "pasteReportedAnswers": 0,
                "suspectedPasteAnswers": 0,
                "suspiciousTypingAnswers": 0,
                "tabSwitchesSum": 0,
                "windowBlurSum": 0,
                "maxWPM": 0.0,
            })

            agg["names"].add(student_name)
            if norm_id:
                agg["normalizedIds"].add(norm_id)
            if session.get("studentEmail"):
                agg["emails"].add(session.get("studentEmail"))

            # Prefer session.clientIp; also collect from accessAttempts
            ip = session.get("clientIp")
            attempt_ips = [a.get("clientIp") for a in (session.get("accessAttempts") or []) if a.get("clientIp")]
            ip_candidates = []
            if ip:
                ip_candidates.append(ip)
            for aip in attempt_ips:
                if aip not in ip_candidates:
                    ip_candidates.append(aip)

            for ip_val in ip_candidates:
                if ip_val and ip_val not in agg["ips_set"]:
                    agg["ips_set"].add(ip_val)
                    agg["ips"].append(ip_val)
                # Map ip to students
                if ip_val:
                    ip_to_students.setdefault(ip_val, set()).add(student_key)

            # User agent
            fp = session.get("browserFingerprint") or {}
            ua = fp.get("userAgent")
            if ua:
                agg["userAgents"].add(ua)

            # Answer-level analysis
            session_flags = {
                "devToolsOpened": 0,
                "pasteReported": 0,
                "suspectedPasteHeuristic": 0,
                "suspiciousTyping": 0,
                "highWPM": 0,
            }
            for ans in answers:
                flags = compute_answer_flags(ans)
                agg["answersCount"] += 1
                agg["devToolsAnswers"] += 1 if flags["devToolsOpened"] else 0
                agg["pasteReportedAnswers"] += 1 if flags["pasteReported"] else 0
                agg["suspectedPasteAnswers"] += 1 if flags["suspectedPasteHeuristic"] else 0
                agg["suspiciousTypingAnswers"] += 1 if flags["suspiciousTypingSpeed"] else 0
                try:
                    if flags["wordsPerMinute"] is not None:
                        agg["maxWPM"] = max(float(agg["maxWPM"] or 0.0), float(flags["wordsPerMinute"]))
                except Exception:
                    pass
                agg["tabSwitchesSum"] += int(flags["tabSwitches"] or 0)
                agg["windowBlurSum"] += int(flags["windowBlurEvents"] or 0)

                session_flags["devToolsOpened"] += 1 if flags["devToolsOpened"] else 0
                session_flags["pasteReported"] += 1 if flags["pasteReported"] else 0
                session_flags["suspectedPasteHeuristic"] += 1 if flags["suspectedPasteHeuristic"] else 0
                session_flags["suspiciousTyping"] += 1 if flags["suspiciousTypingSpeed"] else 0
                session_flags["highWPM"] += 1 if flags["highWPM"] else 0

            entry["_sessionDerivedFlags"] = session_flags
            agg["sessions"].append(entry)

    # Post-process: build issues
    students_overview_rows = []
    session_issues_rows = []
    ip_across_rows = []
    answers_flags_rows = []

    # IP across students
    for ip_val, stu_set in sorted(ip_to_students.items(), key=lambda x: (-len(x[1]), x[0])):
        if ip_val and len(stu_set) > 1:
            sample_students = ", ".join(list(stu_set)[:10])
            ip_across_rows.append([ip_val, len(stu_set), sample_students])

    # Build per-student overview and per-session issues
    for student_key, agg in students.items():
        names_joined = ", ".join(sorted(agg["names"]))
        normalized_ids_joined = ", ".join(sorted(agg["normalizedIds"])) if agg["normalizedIds"] else None
        emails_joined = ", ".join(sorted(agg["emails"])) if agg["emails"] else None
        unique_ips_count = len(agg["ips_set"])
        unique_ua_count = len(agg["userAgents"])
        sample_ips = ", ".join(agg["ips"][:5])
        multiple_ips = unique_ips_count > 1
        multiple_uas = unique_ua_count > 1
        shared_ip = any(len(ip_to_students.get(ip, [])) > 1 for ip in agg["ips_set"])

        students_overview_rows.append([
            student_key,
            names_joined,
            normalized_ids_joined,
            emails_joined,
            len(agg["sessions"]),
            agg["answersCount"],
            unique_ips_count,
            sample_ips,
            unique_ua_count,
            multiple_ips,
            shared_ip,
            multiple_uas,
            agg["devToolsAnswers"],
            agg["pasteReportedAnswers"],
            agg["suspectedPasteAnswers"],
            agg["suspiciousTypingAnswers"],
            agg["tabSwitchesSum"],
            agg["windowBlurSum"],
            agg["maxWPM"],
        ])

        # Determine majority IP for this student (first ip occurrence works as proxy)
        majority_ip = agg["ips"][0] if agg["ips"] else None

        # Per-session issues
        for entry in agg["sessions"]:
            session = entry.get("session", {})
            exam_id = session.get("_id")
            ip = session.get("clientIp") or (session.get("accessAttempts") or [{}])[0].get("clientIp") if (session.get("accessAttempts") or []) else None
            fp = session.get("browserFingerprint") or {}
            ua = fp.get("userAgent")
            flags = entry.get("_sessionDerivedFlags", {})

            issues = []
            if multiple_ips and ip and majority_ip and ip != majority_ip:
                issues.append("IP differs from student's primary IP")
            if shared_ip and ip and len(ip_to_students.get(ip, [])) > 1:
                issues.append("IP shared across multiple students")
            if multiple_uas:
                issues.append("Multiple different user agents across sessions")
            if flags.get("devToolsOpened"):
                issues.append(f"DevTools opened in {flags['devToolsOpened']} answers")
            if flags.get("pasteReported"):
                issues.append(f"Paste reported in {flags['pasteReported']} answers")
            if flags.get("suspectedPasteHeuristic"):
                issues.append(f"Suspected paste (heuristic) in {flags['suspectedPasteHeuristic']} answers")
            if flags.get("suspiciousTyping"):
                issues.append(f"Suspicious typing speed in {flags['suspiciousTyping']} answers")
            if flags.get("highWPM"):
                issues.append(f"Unusually high WPM in {flags['highWPM']} answers")

            if issues:
                session_issues_rows.append([
                    student_key,
                    names_joined,
                    normalized_ids_joined,
                    emails_joined,
                    exam_id,
                    ip,
                    ua,
                    "; ".join(issues)
                ])

            # Collect answer-level flags for this session
            for ans in entry.get("answers", []):
                af = compute_answer_flags(ans)
                reasons = []
                if af["devToolsOpened"]:
                    reasons.append("DevTools")
                if af["pasteReported"]:
                    reasons.append("PasteReported")
                if af["suspectedPasteHeuristic"]:
                    reasons.append("PasteHeuristic")
                if af["suspiciousTypingSpeed"]:
                    reasons.append("SuspiciousTyping")
                if af["highWPM"]:
                    reasons.append("HighWPM")
                if reasons:
                    answers_flags_rows.append([
                        student_key,
                        names_joined,
                        normalized_ids_joined,
                        exam_id,
                        ans.get("questionIndex"),
                        ",".join(reasons),
                        af["wordsPerMinute"],
                        af["copyPasteEvents"],
                        af["typingEventsCount"],
                        af["answerLength"],
                        ans.get("timeSpent"),
                        to_excel_datetime(ans.get("submittedAt")),
                    ])

    return {
        "students_overview_rows": students_overview_rows,
        "session_issues_rows": session_issues_rows,
        "ip_across_rows": ip_across_rows,
        "answers_flags_rows": answers_flags_rows,
    }


def build_summary_sheet(wb, report):
    ws = wb.create_sheet(title="Summary")
    headers = [
        "studentName",
        "sessionsCount",
        "answersCount",
    ]
    ws.append(headers)

    for student in report:
        sessions = student.get("sessions", [])
        answers_count = sum(len(s.get("answers", [])) for s in sessions)
        ws.append([
            student.get("studentName"),
            len(sessions),
            answers_count,
        ])
    autosize_columns(ws)


def build_sessions_sheet(wb, report):
    ws = wb.create_sheet(title="Sessions")
    headers = [
        "studentName",
        "examId",
        "collection",
        "studentEmail",
        "studentId",
        "examTitle",
        "status",
        "startTime",
        "endTime",
        "score",
        "clientIp",
        "fp_userAgent",
        "fp_screenResolution",
        "fp_timezone",
        "accessAttemptsCount",
    ]
    ws.append(headers)

    for student in report:
        for entry in student.get("sessions", []):
            session = entry.get("session", {})
            fp = session.get("browserFingerprint") or {}
            access_attempts = session.get("accessAttempts", [])
            ws.append([
                student.get("studentName"),
                session.get("_id"),
                session.get("collection"),
                session.get("studentEmail"),
                session.get("studentId"),
                session.get("examTitle"),
                session.get("status"),
                to_excel_datetime(session.get("startTime")),
                to_excel_datetime(session.get("endTime")),
                session.get("score"),
                session.get("clientIp"),
                fp.get("userAgent"),
                fp.get("screenResolution"),
                fp.get("timezone") or fp.get("timeZone"),
                len(access_attempts),
            ])
    autosize_columns(ws)


def build_access_attempts_sheet(wb, report):
    ws = wb.create_sheet(title="AccessAttempts")
    headers = [
        "studentName",
        "examId",
        "timestamp",
        "clientIp",
        "success",
        "fingerprint_summary",
    ]
    ws.append(headers)

    for student in report:
        for entry in student.get("sessions", []):
            session = entry.get("session", {})
            exam_id = session.get("_id")
            attempts = session.get("accessAttempts", [])
            for att in attempts:
                fp = att.get("browserFingerprint") or {}
                fp_summary = fp.get("userAgent") or (json.dumps(fp, ensure_ascii=False) if fp else None)
                ws.append([
                    student.get("studentName"),
                    exam_id,
                    to_excel_datetime(att.get("timestamp")),
                    att.get("clientIp"),
                    att.get("success"),
                    fp_summary,
                ])
    autosize_columns(ws)


def build_answers_sheet(wb, report):
    ws = wb.create_sheet(title="Answers")
    headers = [
        "studentName",
        "examId",
        "questionIndex",
        "questionId",
        "difficulty",
        "isCorrect",
        "timeSpent",
        "typingSpeed",
        "isAutoSave",
        "submittedAt",
        # Suspicion/behavior flags
        "suspiciousTypingSpeed",
        "pasteFromExternal",
        "devToolsOpened",
        "tabSwitches",
        "windowBlurEvents",
        "focusScore",
        "totalBackspaces",
        "totalDeletes",
        "editingEfficiency",
        "copyPasteEvents",
        "wordsPerMinute",
        "averageKeyInterval",
    ]
    ws.append(headers)

    for student in report:
        for entry in student.get("sessions", []):
            session = entry.get("session", {})
            exam_id = session.get("_id")
            for ans in entry.get("answers", []):
                ba = ans.get("behaviorAnalytics") or {}
                ws.append([
                    student.get("studentName"),
                    exam_id,
                    ans.get("questionIndex"),
                    ans.get("questionId"),
                    ans.get("difficulty"),
                    ans.get("isCorrect"),
                    ans.get("timeSpent"),
                    ans.get("typingSpeed"),
                    ans.get("isAutoSave"),
                    to_excel_datetime(ans.get("submittedAt")),
                    ba.get("suspiciousTypingSpeed"),
                    ba.get("pasteFromExternal"),
                    ba.get("devToolsOpened"),
                    ba.get("tabSwitches"),
                    ba.get("windowBlurEvents"),
                    ba.get("focusScore"),
                    ba.get("totalBackspaces"),
                    ba.get("totalDeletes"),
                    ba.get("editingEfficiency"),
                    ba.get("copyPasteEvents"),
                    ba.get("wordsPerMinute"),
                    ba.get("averageKeyInterval"),
                ])
    autosize_columns(ws)


def build_students_overview_sheet(wb, rows):
    ws = wb.create_sheet(title="StudentsOverview")
    headers = [
        "studentKey",
        "names",
        "normalizedIds",
        "emails",
        "sessionsCount",
        "answersCount",
        "uniqueIPs",
        "sampleIPs",
        "uniqueUAs",
        "flagMultipleIPs",
        "flagSharedIP",
        "flagMultipleUAs",
        "devToolsAnswers",
        "pasteReportedAnswers",
        "suspectedPasteAnswers",
        "suspiciousTypingAnswers",
        "tabSwitchesSum",
        "windowBlurSum",
        "maxWPM",
    ]
    ws.append(headers)
    for r in rows:
        ws.append(r)
    autosize_columns(ws)


def build_session_issues_sheet(wb, rows):
    ws = wb.create_sheet(title="SessionIssues")
    headers = [
        "studentKey",
        "names",
        "normalizedIds",
        "emails",
        "examId",
        "clientIp",
        "userAgent",
        "issues"
    ]
    ws.append(headers)
    for r in rows:
        ws.append(r)
    autosize_columns(ws)


def build_ip_across_sheet(wb, rows):
    ws = wb.create_sheet(title="IPsAcrossStudents")
    headers = [
        "clientIp",
        "numStudents",
        "studentsSample"
    ]
    ws.append(headers)
    for r in rows:
        ws.append(r)
    autosize_columns(ws)


def build_answers_flags_sheet(wb, rows):
    ws = wb.create_sheet(title="AnswersFlags")
    headers = [
        "studentKey",
        "names",
        "normalizedIds",
        "examId",
        "questionIndex",
        "reasons",
        "wordsPerMinute",
        "copyPasteEvents",
        "typingEventsCount",
        "answerLength",
        "timeSpent",
        "submittedAt",
    ]
    ws.append(headers)
    for r in rows:
        ws.append(r)
    autosize_columns(ws)


def convert_to_excel(input_json_path: str, output_excel_path: str):
    data = load_json(input_json_path)
    report = data.get("report", [])

    analysis = analyze_report(report)

    wb = Workbook()
    # Remove the default sheet
    default_ws = wb.active
    wb.remove(default_ws)

    # Analysis-first sheets for readability
    build_students_overview_sheet(wb, analysis["students_overview_rows"]) 
    build_session_issues_sheet(wb, analysis["session_issues_rows"]) 
    build_ip_across_sheet(wb, analysis["ip_across_rows"]) 
    build_answers_flags_sheet(wb, analysis["answers_flags_rows"]) 

    build_summary_sheet(wb, report)
    build_sessions_sheet(wb, report)
    build_access_attempts_sheet(wb, report)
    build_answers_sheet(wb, report)

    wb.save(output_excel_path)


def main():
    cwd = os.path.dirname(os.path.abspath(__file__))
    input_path = sys.argv[1] if len(sys.argv) > 1 else os.path.join(cwd, 'suspicious_activity_export.json')
    output_path = sys.argv[2] if len(sys.argv) > 2 else os.path.join(cwd, 'suspicious_activity_export.xlsx')

    if not os.path.exists(input_path):
        print(f"Input file not found: {input_path}")
        sys.exit(1)

    convert_to_excel(input_path, output_path)
    print(f"✅ Excel written to: {output_path}")


if __name__ == '__main__':
    main()


